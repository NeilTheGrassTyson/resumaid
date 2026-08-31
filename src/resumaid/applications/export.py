"""Exporting the application history.

CSV is the contract: UTF-8 with BOM and ISO-8601 dates, so a double-click opens it in Excel with
encoding and dates intact. xlsx is a convenience extra.
"""

from __future__ import annotations

import csv
import io
import sqlite3
from pathlib import Path

from resumaid.applications.store import list_applications
from resumaid.util import jload

#: (db column, human-readable header). Flat, one row per application, pivot-table-ready.
COLUMNS: list[tuple[str, str]] = [
    ("company", "Company"),
    ("title", "Position"),
    ("location", "Location"),
    ("submitted_at", "Submitted At"),
    ("submission_channel", "Submitted Via"),
    ("outcome", "Outcome"),
    ("outcome_at", "Outcome Date"),
    ("oa_expected", "OA Expected"),
    ("oa_expectation_confidence", "OA Expectation Confidence"),
    ("oa_received", "OA Received"),
    ("oa_received_at", "OA Received At"),
    ("oa_platform", "OA Platform"),
    ("oa_due_at", "OA Due"),
    ("oa_completed_at", "OA Completed"),
    ("resume_used", "Resume Used"),
    ("fit_score_at_submit", "Fit Score"),
    ("source", "Source"),
    ("apply_url", "Posting URL"),
    ("oa_expectation_evidence", "Why OA Expected"),
    ("notes", "Notes"),
]


def _cell(row: sqlite3.Row, key: str) -> object:
    value = row[key]
    if key == "oa_received":
        # A tri-state: yes, no, or not yet known. Blank is the honest answer for the third.
        return "" if value is None else ("yes" if value else "no")
    if key == "oa_expectation_evidence":
        items = jload(value, []) or []
        return "; ".join(i.get("detail", "") for i in items if isinstance(i, dict))
    if key == "fit_score_at_submit" and value is not None:
        return round(float(value), 1)
    return "" if value is None else value


def to_csv(conn: sqlite3.Connection, **filters: object) -> str:
    rows = list_applications(conn, **filters)  # type: ignore[arg-type]
    buf = io.StringIO(newline="")
    writer = csv.writer(buf, lineterminator="\r\n")
    writer.writerow([header for _, header in COLUMNS])
    for row in rows:
        writer.writerow([_cell(row, key) for key, _ in COLUMNS])
    return buf.getvalue()


def write_csv(conn: sqlite3.Connection, path: Path, **filters: object) -> int:
    """Write the CSV. utf-8-sig gives Excel the BOM it needs to not mangle non-ASCII names."""
    text = to_csv(conn, **filters)
    path.write_text(text, encoding="utf-8-sig", newline="")
    return len(text.splitlines()) - 1


def write_xlsx(conn: sqlite3.Connection, path: Path, **filters: object) -> int:
    """Optional convenience format: real date cells, a frozen header, sized columns."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - depends on the optional extra
        raise RuntimeError(
            "xlsx export needs the 'xlsx' extra: uv pip install -e '.[xlsx]'"
        ) from exc

    rows = list_applications(conn, **filters)  # type: ignore[arg-type]
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    ws.append([header for _, header in COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([_cell(row, key) for key, _ in COLUMNS])
    ws.freeze_panes = "A2"
    for idx, (_, header) in enumerate(COLUMNS, start=1):
        width = max(len(header) + 2, 12)
        ws.column_dimensions[get_column_letter(idx)].width = min(width, 40)
    wb.save(path)
    return len(rows)
